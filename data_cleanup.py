#!/usr/bin/env python3
"""
Data Cleanup and Excel Report Generator

This script automates data cleanup and generates a well-formatted Excel report
from raw CSV or Excel files.

Author: Data Cleanup Script
Version: 1.0
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import sys
import os
from datetime import datetime
import re


class DataCleanupProcessor:
    """Main class for processing and cleaning data files."""
    
    def __init__(self, input_file, output_file=None):
        """
        Initialize the processor with input and output file paths.
        
        Args:
            input_file (str): Path to input CSV or Excel file
            output_file (str): Path to output Excel file (optional)
        """
        self.input_file = input_file
        self.output_file = output_file or self._generate_output_filename()
        self.df = None
        self.original_stats = {}
        self.cleaned_stats = {}
        
    def _generate_output_filename(self):
        """Generate output filename based on input filename."""
        base_name = os.path.splitext(os.path.basename(self.input_file))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_cleaned_{timestamp}.xlsx"
    
    def load_data(self):
        """Load data from CSV or Excel file."""
        try:
            file_ext = os.path.splitext(self.input_file)[1].lower()
            
            if file_ext == '.csv':
                # Try different encodings for CSV files
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        self.df = pd.read_csv(self.input_file, encoding=encoding)
                        print(f"✓ Successfully loaded CSV with {encoding} encoding")
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise Exception("Could not decode CSV file with any supported encoding")
                    
            elif file_ext in ['.xlsx', '.xls']:
                self.df = pd.read_excel(self.input_file)
                print("✓ Successfully loaded Excel file")
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
                
            # Store original statistics
            self.original_stats = self._calculate_stats(self.df)
            print(f"✓ Loaded {len(self.df)} rows and {len(self.df.columns)} columns")
            
        except Exception as e:
            print(f"✗ Error loading file: {str(e)}")
            sys.exit(1)
    
    def _calculate_stats(self, df):
        """Calculate statistics for a dataframe."""
        stats = {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'empty_rows': df.isnull().all(axis=1).sum(),
            'empty_columns': df.isnull().all(axis=0).sum(),
            'duplicate_rows': df.duplicated().sum(),
            'total_empty_cells': df.isnull().sum().sum(),
            'column_info': {}
        }
        
        for col in df.columns:
            stats['column_info'][col] = {
                'unique_values': df[col].nunique(),
                'null_count': df[col].isnull().sum(),
                'data_type': str(df[col].dtype)
            }
        
        return stats
    
    def clean_data(self):
        """Perform comprehensive data cleaning."""
        print("\n🧹 Starting data cleanup process...")
        
        # 1. Remove completely empty rows and columns
        initial_shape = self.df.shape
        self.df = self.df.dropna(how='all')  # Remove empty rows
        self.df = self.df.dropna(axis=1, how='all')  # Remove empty columns
        print(f"✓ Removed empty rows/columns: {initial_shape} → {self.df.shape}")
        
        # 2. Strip whitespace from string columns
        string_cols = self.df.select_dtypes(include=['object']).columns
        for col in string_cols:
            self.df[col] = self.df[col].astype(str).str.strip()
            # Replace 'nan' strings back to NaN
            self.df[col] = self.df[col].replace('nan', np.nan)
        print(f"✓ Stripped whitespace from {len(string_cols)} text columns")
        
        # 3. Standardize date formats
        date_columns = self._detect_date_columns()
        for col in date_columns:
            self.df[col] = self._standardize_dates(self.df[col])
        if date_columns:
            print(f"✓ Standardized date formats in columns: {', '.join(date_columns)}")
        
        # 4. Remove duplicate rows
        duplicates_before = self.df.duplicated().sum()
        self.df = self.df.drop_duplicates()
        if duplicates_before > 0:
            print(f"✓ Removed {duplicates_before} duplicate rows")
        
        # 5. Clean numeric columns
        self._clean_numeric_columns()
        
        # Calculate cleaned statistics
        self.cleaned_stats = self._calculate_stats(self.df)
        print(f"✓ Data cleanup complete! Final shape: {self.df.shape}")
    
    def _detect_date_columns(self):
        """Detect columns that likely contain dates."""
        date_columns = []
        
        for col in self.df.columns:
            # Skip if column is already datetime
            if pd.api.types.is_datetime64_any_dtype(self.df[col]):
                continue
                
            # Check if column name suggests it's a date
            date_keywords = ['date', 'time', 'created', 'updated', 'modified', 'birth', 'dob']
            if any(keyword in col.lower() for keyword in date_keywords):
                date_columns.append(col)
                continue
            
            # Sample non-null values to check if they look like dates
            sample_values = self.df[col].dropna().head(10).astype(str)
            date_like_count = 0
            
            for value in sample_values:
                # Common date patterns
                date_patterns = [
                    r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
                    r'\d{1,2}/\d{1,2}/\d{4}',   # MM/DD/YYYY or DD/MM/YYYY
                    r'\d{1,2}-\d{1,2}-\d{4}',   # MM-DD-YYYY or DD-MM-YYYY
                    r'\d{4}/\d{1,2}/\d{1,2}',   # YYYY/MM/DD
                ]
                
                if any(re.match(pattern, value.strip()) for pattern in date_patterns):
                    date_like_count += 1
            
            # If more than 70% of sampled values look like dates, consider it a date column
            if len(sample_values) > 0 and date_like_count / len(sample_values) > 0.7:
                date_columns.append(col)
        
        return date_columns
    
    def _standardize_dates(self, series):
        """Standardize date formats in a series."""
        try:
            # Try to convert to datetime with various formats
            return pd.to_datetime(series, errors='coerce', infer_datetime_format=True)
        except:
            return series
    
    def _clean_numeric_columns(self):
        """Clean and standardize numeric columns."""
        for col in self.df.columns:
            if self.df[col].dtype == 'object':
                # Try to convert string numbers to numeric
                # Remove common non-numeric characters
                cleaned_series = self.df[col].astype(str).str.replace(r'[,$%]', '', regex=True)
                cleaned_series = cleaned_series.replace('nan', np.nan)
                
                # Try to convert to numeric
                numeric_series = pd.to_numeric(cleaned_series, errors='coerce')
                
                # If more than 50% of non-null values are numeric, convert the column
                non_null_original = self.df[col].notna().sum()
                non_null_numeric = numeric_series.notna().sum()
                
                if non_null_original > 0 and non_null_numeric / non_null_original > 0.5:
                    self.df[col] = numeric_series
                    print(f"✓ Converted column '{col}' to numeric")
    
    def generate_summary_stats(self):
        """Generate comprehensive summary statistics."""
        summary_data = []
        
        # Overall statistics
        summary_data.append(['Metric', 'Original', 'Cleaned', 'Change'])
        summary_data.append(['Total Rows', self.original_stats['total_rows'], 
                           self.cleaned_stats['total_rows'],
                           self.cleaned_stats['total_rows'] - self.original_stats['total_rows']])
        summary_data.append(['Total Columns', self.original_stats['total_columns'],
                           self.cleaned_stats['total_columns'],
                           self.cleaned_stats['total_columns'] - self.original_stats['total_columns']])
        summary_data.append(['Empty Rows Removed', self.original_stats['empty_rows'], 0,
                           -self.original_stats['empty_rows']])
        summary_data.append(['Empty Columns Removed', self.original_stats['empty_columns'], 0,
                           -self.original_stats['empty_columns']])
        summary_data.append(['Duplicate Rows Removed', self.original_stats['duplicate_rows'], 0,
                           -self.original_stats['duplicate_rows']])
        summary_data.append(['Total Empty Cells', self.original_stats['total_empty_cells'],
                           self.cleaned_stats['total_empty_cells'],
                           self.cleaned_stats['total_empty_cells'] - self.original_stats['total_empty_cells']])
        
        # Column-wise statistics
        summary_data.append(['', '', '', ''])  # Empty row
        summary_data.append(['Column Statistics', '', '', ''])
        summary_data.append(['Column Name', 'Unique Values', 'Null Count', 'Data Type'])
        
        for col in self.df.columns:
            col_stats = self.cleaned_stats['column_info'][col]
            summary_data.append([col, col_stats['unique_values'], 
                               col_stats['null_count'], col_stats['data_type']])
        
        return summary_data
    
    def save_excel_report(self):
        """Save cleaned data and summary to Excel with formatting."""
        print(f"\n📊 Generating Excel report: {self.output_file}")
        
        try:
            # Create Excel writer object
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                # Write cleaned data to 'Cleaned Data' sheet
                self.df.to_excel(writer, sheet_name='Cleaned Data', index=False)
                
                # Write summary statistics to 'Summary' sheet
                summary_data = self.generate_summary_stats()
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Apply formatting
            self._apply_excel_formatting()
            
            print(f"✓ Excel report saved successfully: {self.output_file}")
            
        except Exception as e:
            print(f"✗ Error saving Excel report: {str(e)}")
            sys.exit(1)
    
    def _apply_excel_formatting(self):
        """Apply conditional formatting and styling to Excel file."""
        try:
            wb = load_workbook(self.output_file)
            
            # Format 'Cleaned Data' sheet
            if 'Cleaned Data' in wb.sheetnames:
                ws_data = wb['Cleaned Data']
                
                # Header formatting
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in ws_data[1]:  # First row (headers)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Highlight empty cells in light red
                empty_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                for row in ws_data.iter_rows(min_row=2):  # Skip header
                    for cell in row:
                        if cell.value is None or str(cell.value).strip() == '':
                            cell.fill = empty_fill
            
            # Format 'Summary' sheet
            if 'Summary' in wb.sheetnames:
                ws_summary = wb['Summary']
                
                # Header formatting for summary
                summary_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                summary_header_font = Font(color='FFFFFF', bold=True)
                
                # Format specific header rows
                header_rows = [1, 9, 10]  # Adjust based on summary structure
                for row_num in header_rows:
                    try:
                        for cell in ws_summary[row_num]:
                            if cell.value:
                                cell.fill = summary_header_fill
                                cell.font = summary_header_font
                    except:
                        pass
            
            wb.save(self.output_file)
            print("✓ Applied Excel formatting successfully")
            
        except Exception as e:
            print(f"⚠ Warning: Could not apply formatting: {str(e)}")


def main():
    """Main function to run the data cleanup script."""
    parser = argparse.ArgumentParser(description='Data Cleanup and Excel Report Generator')
    parser.add_argument('input_file', help='Path to input CSV or Excel file')
    parser.add_argument('-o', '--output', help='Path to output Excel file (optional)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not os.path.exists(args.input_file):
        print(f"✗ Error: Input file '{args.input_file}' not found")
        sys.exit(1)
    
    print("🚀 Data Cleanup and Excel Report Generator")
    print("=" * 50)
    print(f"Input file: {args.input_file}")
    
    try:
        # Initialize processor
        processor = DataCleanupProcessor(args.input_file, args.output)
        
        # Process data
        processor.load_data()
        processor.clean_data()
        processor.save_excel_report()
        
        print("\n" + "=" * 50)
        print("🎉 Process completed successfully!")
        print(f"📁 Output file: {processor.output_file}")
        
    except KeyboardInterrupt:
        print("\n\n⚠ Process interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Unexpected error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()